import openpyxl


def getRowCount(file,sheetName):
    Workbook=openpyxl.load_workbook(file)
    sheet=Workbook[sheetName]
    return (sheet.max_row)

def getColumnCount(file,sheetName):
    Workbook=openpyxl.load_workbook(file)
    sheet=Workbook[sheetName]
    return (sheet.max_column)


def readData(file,sheetName,rownum,columnno):
    Workbook=openpyxl.load_workbook(file)
    sheet=Workbook[sheetName]
    return sheet.cell(row=rownum,column=columnno).value


def writeData(file,sheetName,rownum,columnno,data):
    Workbook=openpyxl.load_workbook(file)
    sheet=Workbook[sheetName]
    sheet.cell(row=rownum,column=columnno).value=data
    Workbook.save(file)